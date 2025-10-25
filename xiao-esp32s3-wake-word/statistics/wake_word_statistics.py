#!/usr/bin/env python3
"""
Wake Word Detection Statistics Tool
Monitors ESP32-S3 serial output, collects detection statistics, and exports to Excel.
"""

import serial
import serial.tools.list_ports
import re
import json
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Optional, Tuple
import os
import sys
import time
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    import pandas as pd
    from colorama import init, Fore, Style
except ImportError as e:
    print(f"Error: Missing required package. Please install: pip install -r requirements.txt")
    print(f"Missing: {e}")
    sys.exit(1)

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# Constants from esp32s3_standalone.ino
CLASS_LABELS = ["lehitraoot", "shalom", "background", "unknown"]
SERIAL_BAUD_RATE = 921600
CONFIDENCE_THRESHOLD = 0.7

class Detection:
    """Represents a single wake word detection."""
    def __init__(self, detection_num: int, class_name: str, confidence: float, 
                 confidences: List[float], timestamp: datetime):
        self.detection_num = detection_num
        self.class_name = class_name
        self.confidence = confidence
        self.confidences = confidences  # All class confidences
        self.timestamp = timestamp
        self.is_false_positive = False

class Session:
    """Represents a test session."""
    def __init__(self, session_id: int, wake_word: str, distance: str, expected_count: int):
        self.session_id = session_id
        self.wake_word = wake_word
        self.distance = distance
        self.expected_count = expected_count
        self.start_time = datetime.now()
        self.end_time = None
        self.detections: List[Detection] = []
        self.false_positives: List[int] = []  # Detection numbers that are false positives
    
    def add_detection(self, detection: Detection):
        self.detections.append(detection)
    
    def mark_false_positive(self, detection_num: int):
        """Mark a detection as a false positive."""
        if detection_num in [d.detection_num for d in self.detections]:
            self.false_positives.append(detection_num)
            for detection in self.detections:
                if detection.detection_num == detection_num:
                    detection.is_false_positive = True
    
    def end_session(self):
        self.end_time = datetime.now()
    
    @property
    def actual_detections(self) -> int:
        return len([d for d in self.detections if d.class_name == self.wake_word])
    
    @property
    def false_positive_count(self) -> int:
        return len(self.false_positives)
    
    @property
    def true_positives(self) -> int:
        return self.actual_detections - self.false_positive_count
    
    @property
    def detection_rate(self) -> float:
        if self.expected_count == 0:
            return 0.0
        return (self.true_positives / self.expected_count) * 100
    
    @property
    def avg_confidence(self) -> float:
        if not self.detections:
            return 0.0
        relevant = [d for d in self.detections if d.class_name == self.wake_word and not d.is_false_positive]
        if not relevant:
            return 0.0
        return sum(d.confidence for d in relevant) / len(relevant)

class StatisticsCollector:
    """Main class for collecting and managing statistics."""
    
    def __init__(self, results_dir: str = "results"):
        self.sessions: List[Session] = []
        self.current_session: Optional[Session] = None
        self.results_dir = Path(results_dir)
        self.results_dir.mkdir(exist_ok=True)
        self.detection_counter = 0
        
    def start_session(self, session_id: int, wake_word: str, distance: str, expected_count: int):
        """Start a new test session."""
        self.current_session = Session(session_id, wake_word, distance, expected_count)
        print(f"\n{Fore.GREEN}[Session {session_id}] Started: {wake_word} at {distance} distance")
        print(f"Expected count: {expected_count}")
        print(f"Press 'f' to mark last detection as false positive")
        print(f"Press 'q' to end session\n")
    
    def end_session(self):
        """End the current session."""
        if self.current_session:
            self.current_session.end_session()
            self.sessions.append(self.current_session)
            self._print_session_summary(self.current_session)
            self.current_session = None
    
    def add_detection(self, class_name: str, confidence: float, confidences: List[float]):
        """Add a detection to the current session."""
        if not self.current_session:
            return
        
        self.detection_counter += 1
        detection = Detection(
            detection_num=self.detection_counter,
            class_name=class_name,
            confidence=confidence,
            confidences=confidences,
            timestamp=datetime.now()
        )
        self.current_session.add_detection(detection)
        
        # Print detection with color coding
        if class_name == self.current_session.wake_word:
            color = Fore.GREEN
            status = "CORRECT"
        else:
            color = Fore.YELLOW
            status = "WRONG CLASS"
        
        print(f"{color}[#{self.detection_counter}] {class_name} - Confidence: {confidence:.3f} [{status}]")
    
    def mark_last_false_positive(self):
        """Mark the last detection as a false positive."""
        if self.current_session and self.detection_counter > 0:
            self.current_session.mark_false_positive(self.detection_counter)
            print(f"{Fore.RED}Marked detection #{self.detection_counter} as FALSE POSITIVE")
    
    def _print_session_summary(self, session: Session):
        """Print a summary of the session."""
        print(f"\n{Fore.CYAN}{'='*60}")
        print(f"Session {session.session_id} Summary")
        print(f"{'='*60}{Style.RESET_ALL}")
        print(f"Wake word: {session.wake_word}")
        print(f"Distance: {session.distance}")
        print(f"Expected: {session.expected_count}")
        print(f"Actual detections: {session.actual_detections}")
        print(f"False positives: {session.false_positive_count}")
        print(f"True positives: {session.true_positives}")
        print(f"Detection rate: {session.detection_rate:.1f}%")
        print(f"Average confidence: {session.avg_confidence:.3f}")
        print(f"{'='*60}\n")
    
    def export_to_excel(self) -> str:
        """Export all sessions to an Excel file."""
        if not self.sessions:
            print("No sessions to export.")
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.results_dir / f"wake_word_stats_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        self._create_summary_sheet(wb)
        
        # Create Sessions sheet
        self._create_sessions_sheet(wb)
        
        # Create Detections sheet
        self._create_detections_sheet(wb)
        
        wb.save(filename)
        return str(filename)

    def _create_summary_sheet(self, wb: Workbook):
        """Create the Summary sheet with aggregated statistics."""
        ws = wb.create_sheet("Summary")
        
        # Headers
        headers = ["Wake Word", "Distance", "Sessions", "Total Expected", 
                  "Total Detected", "False Positives", "True Positives", 
                  "Detection Rate (%)", "Avg Confidence"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Aggregate data by wake word and distance
        summary_data = defaultdict(lambda: {
            'sessions': 0,
            'total_expected': 0,
            'total_detected': 0,
            'false_positives': 0,
            'true_positives': 0,
            'confidences': []
        })
        
        for session in self.sessions:
            key = (session.wake_word, session.distance)
            summary_data[key]['sessions'] += 1
            summary_data[key]['total_expected'] += session.expected_count
            summary_data[key]['total_detected'] += session.actual_detections
            summary_data[key]['false_positives'] += session.false_positive_count
            summary_data[key]['true_positives'] += session.true_positives
            if session.avg_confidence > 0:
                summary_data[key]['confidences'].append(session.avg_confidence)
        
        # Write data
        row = 2
        for (wake_word, distance), data in sorted(summary_data.items()):
            ws.cell(row=row, column=1, value=wake_word)
            ws.cell(row=row, column=2, value=distance)
            ws.cell(row=row, column=3, value=data['sessions'])
            ws.cell(row=row, column=4, value=data['total_expected'])
            ws.cell(row=row, column=5, value=data['total_detected'])
            ws.cell(row=row, column=6, value=data['false_positives'])
            ws.cell(row=row, column=7, value=data['true_positives'])
            
            detection_rate = (data['true_positives'] / data['total_expected'] * 100) if data['total_expected'] > 0 else 0
            ws.cell(row=row, column=8, value=f"{detection_rate:.1f}")
            
            avg_conf = sum(data['confidences']) / len(data['confidences']) if data['confidences'] else 0
            ws.cell(row=row, column=9, value=f"{avg_conf:.3f}")
            
            row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_sessions_sheet(self, wb: Workbook):
        """Create the Sessions sheet with detailed session data."""
        ws = wb.create_sheet("Sessions")
        
        headers = ["Session ID", "Wake Word", "Distance", "Expected", 
                  "Detected", "False Positives", "True Positives", 
                  "Detection Rate (%)", "Avg Confidence", "Start Time", "Duration"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, session in enumerate(self.sessions, start=2):
            duration = (session.end_time - session.start_time).total_seconds() if session.end_time else 0
            
            ws.cell(row=row_idx, column=1, value=session.session_id)
            ws.cell(row=row_idx, column=2, value=session.wake_word)
            ws.cell(row=row_idx, column=3, value=session.distance)
            ws.cell(row=row_idx, column=4, value=session.expected_count)
            ws.cell(row=row_idx, column=5, value=session.actual_detections)
            ws.cell(row=row_idx, column=6, value=session.false_positive_count)
            ws.cell(row=row_idx, column=7, value=session.true_positives)
            ws.cell(row=row_idx, column=8, value=f"{session.detection_rate:.1f}")
            ws.cell(row=row_idx, column=9, value=f"{session.avg_confidence:.3f}")
            ws.cell(row=row_idx, column=10, value=session.start_time.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=11, value=f"{duration:.1f}s")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_detections_sheet(self, wb: Workbook):
        """Create the Detections sheet with individual detection records."""
        ws = wb.create_sheet("Detections")
        
        headers = ["Detection #", "Session ID", "Wake Word", "Distance", 
                  "Class", "Confidence", "Lehitraoot", "Shalom", 
                  "Background", "Unknown", "False Positive", "Timestamp"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for session in self.sessions:
            for detection in session.detections:
                ws.cell(row=row, column=1, value=detection.detection_num)
                ws.cell(row=row, column=2, value=session.session_id)
                ws.cell(row=row, column=3, value=session.wake_word)
                ws.cell(row=row, column=4, value=session.distance)
                ws.cell(row=row, column=5, value=detection.class_name)
                ws.cell(row=row, column=6, value=f"{detection.confidence:.3f}")
                ws.cell(row=row, column=7, value=f"{detection.confidences[0]:.3f}")
                ws.cell(row=row, column=8, value=f"{detection.confidences[1]:.3f}")
                ws.cell(row=row, column=9, value=f"{detection.confidences[2]:.3f}")
                ws.cell(row=row, column=10, value=f"{detection.confidences[3]:.3f}")
                ws.cell(row=row, column=11, value="Yes" if detection.is_false_positive else "No")
                ws.cell(row=row, column=12, value=detection.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

class SerialMonitor:
    """Handles serial communication with ESP32-S3."""
    
    def __init__(self, port: str, baud_rate: int = SERIAL_BAUD_RATE):
        self.port = port
        self.baud_rate = baud_rate
        self.serial_connection = None
        
    def connect(self) -> bool:
        """Connect to the serial port."""
        try:
            self.serial_connection = serial.Serial(
                port=self.port,
                baudrate=self.baud_rate,
                timeout=1
            )
            print(f"{Fore.GREEN}Connected to {self.port} at {self.baud_rate} baud{Style.RESET_ALL}")
            return True
        except Exception as e:
            print(f"{Fore.RED}Error connecting to {self.port}: {e}{Style.RESET_ALL}")
            return False
    
    def disconnect(self):
        """Disconnect from the serial port."""
        if self.serial_connection and self.serial_connection.is_open:
            self.serial_connection.close()
            print(f"{Fore.YELLOW}Disconnected from {self.port}{Style.RESET_ALL}")
    
    def read_line(self) -> Optional[str]:
        """Read a line from the serial port."""
        if self.serial_connection and self.serial_connection.is_open:
            try:
                line = self.serial_connection.readline().decode('utf-8', errors='ignore').strip()
                return line if line else None
            except Exception as e:
                print(f"{Fore.RED}Serial read error: {e}{Style.RESET_ALL}")
                return None
        return None
    
    def flush_input_buffer(self):
        """Flush the serial input buffer to discard old data."""
        if self.serial_connection and self.serial_connection.is_open:
            self.serial_connection.reset_input_buffer()

def parse_detection_message(line: str, parsing_state: dict) -> Tuple[Optional[dict], dict]:
    """
    Parse a detection message from the ESP32-S3.
    
    Expected format:
    *** WAKE WORD DETECTED #X ***
    Class: <label>, Confidence: 0.XXX (threshold: 0.700)
    All confidences: [0.XXX, 0.XXX, 0.XXX, 0.XXX]
    
    Returns: (detection_dict or None, updated parsing_state)
    """
    detection_pattern = r'\*\*\* WAKE WORD DETECTED #(\d+) \*\*\*'
    detection_match = re.search(detection_pattern, line)
    
    if detection_match:
        # Start parsing a new detection
        parsing_state['waiting_for_class'] = True
        parsing_state['waiting_for_confidences'] = False
        return None, parsing_state
    
    # Match class and confidence line
    if parsing_state.get('waiting_for_class', False):
        class_pattern = r'Class: (\w+), Confidence: ([\d.]+)'
        class_match = re.search(class_pattern, line)
        
        if class_match:
            parsing_state['class_name'] = class_match.group(1)
            parsing_state['confidence'] = float(class_match.group(2))
            parsing_state['waiting_for_class'] = False
            parsing_state['waiting_for_confidences'] = True
            return None, parsing_state
    
    # Match confidences array
    if parsing_state.get('waiting_for_confidences', False):
        confidences_pattern = r'All confidences: \[([\d., ]+)\]'
        confidences_match = re.search(confidences_pattern, line)
        
        if confidences_match:
            confidences_str = confidences_match.group(1)
            confidences = [float(x.strip()) for x in confidences_str.split(',')]
            
            # Complete detection
            detection = {
                'class_name': parsing_state['class_name'],
                'confidence': parsing_state['confidence'],
                'confidences': confidences
            }
            
            # Reset parsing state
            parsing_state = {}
            return detection, parsing_state
    
    return None, parsing_state

def find_esp32_port() -> Optional[str]:
    """Find and select the ESP32-S3 COM port with a menu."""
    import serial.tools.list_ports
    
    print("Scanning for ESP32...")
    ports = serial.tools.list_ports.comports()
    
    # Look for common ESP32 identifiers
    esp32_keywords = ['ch340', 'ch341', 'cp210', 'ftdi', 'usb-to-serial', 'uart', 'silicon labs']
    
    for port in ports:
        # Look for common ESP32 USB chip identifiers
        if any(x in port.description.lower() for x in esp32_keywords):
            print(f"Found potential ESP32 at: {port.device} ({port.description})")
            return port.device
    
    # Show all available ports
    if ports:
        print("\nAvailable ports:")
        for i, port in enumerate(ports):
            print(f"  [{i}] {port.device} - {port.description}")
        
        choice = input("\nEnter port number or full port name (e.g., COM3): ").strip()
        if choice.isdigit() and int(choice) < len(ports):
            return ports[int(choice)].device
        else:
            return choice
    
    return None

def get_user_input(prompt: str, valid_options: List[str] = None) -> str:
    """Get user input with optional validation."""
    while True:
        response = input(prompt).strip().lower()
        if valid_options is None or response in valid_options:
            return response
        print(f"{Fore.RED}Invalid input. Please enter one of: {', '.join(valid_options)}{Style.RESET_ALL}")

def main():
    """Main entry point for the statistics tool."""
    print(f"{Fore.CYAN}{Style.BRIGHT}Wake Word Detection Statistics Tool{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'='*60}{Style.RESET_ALL}\n")
    
    # Initialize collector
    collector = StatisticsCollector()
    
    # Find and select COM port with menu
    port = find_esp32_port()
    
    if not port:
        print("\nNo ESP32 found. Please enter the COM port manually.")
        port = input("Port (e.g., COM3): ").strip()
    
    if not port:
        print("Error: No port specified")
        return
    
    # Connect to ESP32
    monitor = SerialMonitor(port)
    if not monitor.connect():
        print("Failed to connect to ESP32. Exiting.")
        return
    
    try:
        session_id = 1
        
        while True:
            # Get session parameters
            print(f"\n{Fore.CYAN}{'='*60}{Style.RESET_ALL}")
            print(f"{Fore.CYAN}Session {session_id} Configuration{Style.RESET_ALL}")
            print(f"{Fore.CYAN}{'='*60}{Style.RESET_ALL}\n")
            
            # Distance selection with menu
            print("Enter distance category: [1] close / [2] far")
            distance_choice = get_user_input("Selection: ", ['1', '2'])
            distance = 'close' if distance_choice == '1' else 'far'
            
            # Wake word selection with menu
            print("\nWhich wake word are you testing? [1] lehitraoot / [2] shalom / [3] both")
            wake_word_choice = get_user_input("Selection: ", ['1', '2', '3'])
            wake_word_map = {'1': 'lehitraoot', '2': 'shalom', '3': 'both'}
            wake_word = wake_word_map[wake_word_choice]
            expected_count = int(get_user_input("How many times will you say it? "))
            
            # Start session
            collector.start_session(session_id, wake_word, distance, expected_count)
            
            # Flush any buffered serial data from before the session started
            monitor.flush_input_buffer()
            print(f"{Fore.CYAN}Serial buffer cleared. Ready to monitor.{Style.RESET_ALL}\n")
            
            # Monitoring loop
            parsing_state = {}
            
            print(f"{Fore.YELLOW}Monitoring... (Press 'f' for false positive, 'q' to end session){Style.RESET_ALL}\n")
            
            # Set up silent keypress detection for Windows
            session_active = True
            
            try:
                import sys
                if sys.platform == 'win32':
                    from msvcrt import kbhit, getch
                    use_windows_input = True
                else:
                    use_windows_input = False
            except ImportError:
                use_windows_input = False
            
            while session_active:
                # Check for silent keypress on Windows
                if use_windows_input:
                    try:
                        if kbhit():
                            key = getch()
                            if isinstance(key, bytes):
                                key = key.decode('utf-8').lower()
                            else:
                                key = key.lower()
                            
                            if key == 'q':
                                print(f"\n{Fore.YELLOW}Ending session...{Style.RESET_ALL}")
                                session_active = False
                                break
                            elif key == 'f':
                                collector.mark_last_false_positive()
                    except:
                        pass
                
                # Read serial line
                line = monitor.read_line()
                
                if line:
                    # Parse detection message
                    detection, parsing_state = parse_detection_message(line, parsing_state)
                    
                    if detection:
                        # Complete detection received
                        collector.add_detection(
                            detection['class_name'],
                            detection['confidence'],
                            detection['confidences']
                        )
                
                # Check for 'q' to quit
                if not session_active:
                    break
                
                # Small delay to prevent overwhelming the CPU
                time.sleep(0.01)
            
            # End session
            collector.end_session()
            
            # Ask for another session
            continue_session = get_user_input("\nStart another session? (y/n): ", ['y', 'n'])
            if continue_session == 'n':
                break
            
            session_id += 1
        
        # Export results
        print(f"\n{Fore.CYAN}Exporting results to Excel...{Style.RESET_ALL}")
        filename = collector.export_to_excel()
        if filename:
            print(f"{Fore.GREEN}Results saved to: {filename}{Style.RESET_ALL}")
        
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}Interrupted by user. Saving data...{Style.RESET_ALL}")
        if collector.sessions:
            filename = collector.export_to_excel()
            if filename:
                print(f"{Fore.GREEN}Results saved to: {filename}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}Error: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
    finally:
        monitor.disconnect()
        print(f"\n{Fore.CYAN}Exiting. Goodbye!{Style.RESET_ALL}")

if __name__ == "__main__":
    main()
